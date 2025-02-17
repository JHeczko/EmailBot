import string

from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


def edit_excel(workbook,mode,i_mama, i_mail, i_r1, i_r2, i_r3, opt_indexes):
    '''
    basicly a heart of a program which preprocess a excel files(processing logic here). Also it can work in 3 modes:

    * NAME SURNAME(mode 0): first name and then surname
    * SURNAME NAME(mode 1): first surname and then name

    :param workbook: a workbook to be processed
    :param mode: how to process names and surnames, explained at the top
    :param i_mama: index of name and surname of Mum inside a excel table
    :param i_mail: index of mail
    :param i_r1: index of first interval of payment
    :param i_r2: index of second interval of payment
    :param i_r3: index of third interval of payment
    :param opt_indexes: other interval of payments to be processed
    :return: processed and nicely colored workbook
    '''
    newWorkBook = Workbook()
    newSheet = newWorkBook.active
    sheet = workbook.active  # Pobieramy aktywny arkusz

    # idexes of all money column
    r_indexes = [i_r1, i_r2, i_r3]
    r_indexes.extend(opt_indexes)

    # hashmaps based on name and surname of a person(expecting mother or father not kids)
    hashImieRn = [{} for _ in r_indexes]
    hashImieALL = {}

    hashImieMail = {}
    hashImieIlosc = {}

    # =-=-=-=-=-=-=-=-=SORTING DATA FROM OLD WORKBOOK=-=-=-=-=-=-=-=-=
    for row in sheet.iter_rows(min_row=2,min_col=0, values_only=True):
        mama = row[i_mama]
        if row[i_mama] is None:
            mama = 'nieznany_rodzic'
        elif row[i_mama].translate(str.maketrans('','', string.whitespace)) == '':
            mama = 'nieznany_rodzic'

        mail = row[i_mail].translate(str.maketrans('','', string.whitespace)) if row[i_mail] is not None else ''

        moneyALL = 0
        # iterate through every hashmap for every nterval and add a specific value to a mum index
        for i_r, hashImie in zip(r_indexes, hashImieRn):
            moneyRn = int(row[i_r]) if row[i_r] is not None else 0
            moneyALL += moneyRn
            if mama in hashImie:
                hashImie[mama] += moneyRn
            else:
                hashImie[mama] = moneyRn

        if mama in hashImieALL:
            hashImieALL[mama] += moneyALL
        else:
            hashImieALL[mama] = moneyALL

        # adding a child to a specific mama
        if mama in hashImieIlosc:
            hashImieIlosc[mama] += 1
        else:
            hashImieIlosc[mama] = 1

        # adding a mail to a specific mama(cuz all mail should be the same so it doesnt really matter here)
        if mama not in hashImieMail:
            hashImieMail[mama] = mail

    # descending sorting
    hashImieALL = dict(sorted(hashImieALL.items(), key=lambda x: x[1], reverse=False))

    # =-=-=-=-=-=-=-=-=STYLING AND FILLING NEW WORKBOOK=-=-=-=-=-=-=-=-=
    # ============Styling============
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='left', vertical='center')
    light_gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')


    # ============Header creation============
    headers = ['Zwrot','Nazwisko', 'Mail1','Mail2', 'Ilość Dzieci']
    headers.extend([f"R{i}" for i,_ in enumerate(r_indexes,1)])
    headers.append('ALL')
    for col_num, header in enumerate(headers, start=1):
        cell = newSheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = PatternFill(start_color='00969696', end_color='00969696', fill_type='solid')

    # ============Filling data============
    row = 2
    for mama in hashImieALL.keys():
        nazwiskoMama = mama.split()

        # name processing
        if mode == 0:
            nazwiskoMama = nazwiskoMama[1:] if len(nazwiskoMama) != 1 else nazwiskoMama[0] + " nieznane_nazwisko"
            nazwiskoMama = ''.join(nazwiskoMama)
        elif mode == 1:
            nazwiskoMama.reverse()
            nazwiskoMama = nazwiskoMama[1:] if len(nazwiskoMama) != 1 else nazwiskoMama[0] + " nieznane_nazwisko"
            nazwiskoMama = ''.join(nazwiskoMama)

        # mail processing
        mails = hashImieMail[mama].split(';')

        # removing all empty string values from split
        for mail in mails:
            if mail.translate(str.maketrans('','', string.whitespace)) == '':
                mails.remove(mail)

        # only taking two mails here, later i can add adding additional column autoamicly if 2 or more mails are detected for now staying with fixed 2 mail ammount
        mailMama = mails[0] if len(mails) > 0 else ''
        mailMama2 = mails[1] if len(mails) > 1 else ''

        newSheet.cell(row=row, column=1, value="Szanowna Pani")
        newSheet.cell(row=row, column=2, value=nazwiskoMama if mama!="nieznany_rodzic" else "Nie znaleziono mamy!")

        newSheet.cell(row=row, column=3, value= mailMama if mama!="nieznany_rodzic" else "Uzupelnij dane o imie oraz nazwisko mamy")
        newSheet.cell(row=row, column=4, value=mailMama2 if mama != "nieznany_rodzic" else "Uzupelnij dane o imie oraz nazwisko mamy")

        newSheet.cell(row=row, column=5, value=hashImieIlosc[mama])

        col_all = 0
        for col,hashImie in enumerate(hashImieRn,6):
            newSheet.cell(row=row, column=col, value=hashImie[mama]).number_format = '#,##0.00 "zł"'
            col_all = col + 1
        newSheet.cell(row=row, column=col_all, value=hashImieALL[mama]).number_format = '#,##0.00 "zł"'
        row += 1

    # ============Styling============
    row = 2
    for mama in hashImieMail.keys():
        for i in range(1,len(headers)+1):
            if row % 2 != 0:
                newSheet.cell(row=row, column=i).fill = light_gray_fill
        row += 1

    row = 2
    for mama in hashImieMail.keys():
        for i in range(1,len(headers)+1):
            newSheet.cell(row=row, column=i).alignment = cell_alignment
        row += 1

    # Adjust column widths
    for col in newSheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = (max_length + 10)  # Add some extra space
        newSheet.column_dimensions[column].width = adjusted_width
    return newWorkBook

if __name__ == "__main__":
    workbook = load_workbook("tests/test_parsing.xlsx")
    edited_workbook = edit_excel(workbook= workbook, mode=0,i_mama=4, i_mail=6, i_r1=8, i_r2=9, i_r3=10, opt_indexes=[11,12,13])
    edited_workbook.save("./testy_parser.xlsx")